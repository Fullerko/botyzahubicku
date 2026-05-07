import csv
import io
import re
import unicodedata
from datetime import datetime

from . import db
from .models import Product, ProductVariant
from .woocommerce_api import update_supplier_variant_in_woocommerce


HEADER_ALIASES = {
    'product_name': {
        'product_name', 'product', 'name', 'nazev', 'nazev_produktu', 'název', 'název produktu',
        'produkt', 'model', 'item_name', 'title'
    },
    'internal_sku': {
        'internal_sku', 'bzh_sku', 'interni_sku', 'interní sku', 'sku_webu', 'web_sku', 'sku eshopu'
    },
    'size': {
        'size', 'velikost', 'shoe_size', 'eu_size', 'eu velikost', 'rozmer', 'rozměr'
    },
    'color': {
        'color', 'colour', 'barva', 'color_supplier', 'supplier_color', 'barva dodavatele',
        'dodavatelska_barva', 'dodavatelská barva'
    },
    'supplier_sku': {
        'supplier_sku', 'sku', 'dodavatelske_sku', 'dodavatelské sku', 'sku dodavatele',
        'variant_sku', 'sku_varianty', 'seller_sku'
    },
    'supplier_color': {
        'supplier_color', 'barva_dodavatele', 'barva dodavatele', 'supplier colour', 'supplier_color_name'
    },
    'supplier_product_code': {
        'supplier_product_code', 'product_code', 'kod_produktu', 'kód produktu', 'kod dodavatele',
        'kód dodavatele', 'supplier_code', 'spu', 'item_code'
    },
    'supplier_ean': {
        'ean', 'barcode', 'čárový kód', 'carovy_kod', 'supplier_ean', 'gtin'
    },
}

COLOR_CANONICAL = {
    # Czech -> English/common key
    'cerna': 'black', 'cerny': 'black', 'cerne': 'black', 'black': 'black', 'blk': 'black', 'heise': 'black', '黑色': 'black',
    'bila': 'white', 'bily': 'white', 'bile': 'white', 'white': 'white', 'wht': 'white', 'baise': 'white', '白色': 'white',
    'seda': 'gray', 'sediva': 'gray', 'sedy': 'gray', 'grey': 'gray', 'gray': 'gray', 'gry': 'gray', 'huise': 'gray', '灰色': 'gray',
    'cervena': 'red', 'cerveny': 'red', 'red': 'red', 'hongse': 'red', '红色': 'red',
    'modra': 'blue', 'modry': 'blue', 'blue': 'blue', 'lanse': 'blue', '蓝色': 'blue',
    'zelena': 'green', 'zeleny': 'green', 'green': 'green', 'lvse': 'green', '绿色': 'green',
    'zluta': 'yellow', 'zluty': 'yellow', 'yellow': 'yellow', 'huangse': 'yellow', '黄色': 'yellow',
    'ruzova': 'pink', 'ruzovy': 'pink', 'pink': 'pink', 'fense': 'pink', '粉色': 'pink',
    'fialova': 'purple', 'fialovy': 'purple', 'purple': 'purple', 'zise': 'purple', '紫色': 'purple',
    'hneda': 'brown', 'hnedy': 'brown', 'brown': 'brown', 'zongse': 'brown', '棕色': 'brown',
    'bezova': 'beige', 'bezovy': 'beige', 'beige': 'beige', 'm色': 'beige', '米色': 'beige',
    'oranzova': 'orange', 'oranzovy': 'orange', 'orange': 'orange', '橙色': 'orange',
    'stribrna': 'silver', 'stribrny': 'silver', 'silver': 'silver', '银色': 'silver',
    'zlata': 'gold', 'zlaty': 'gold', 'gold': 'gold', '金色': 'gold',
    'khaki': 'khaki',
}


def _norm(value):
    value = str(value or '').strip().lower()
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^a-z0-9]+', ' ', value).strip()
    return value


def _compact(value):
    return re.sub(r'\s+', '', _norm(value))


def _color_key(value):
    raw = str(value or '').strip()
    if raw in COLOR_CANONICAL:
        return COLOR_CANONICAL[raw]
    compact = _compact(raw)
    return COLOR_CANONICAL.get(compact, compact)


def _header_key(header):
    normalized = _norm(header)
    compact = normalized.replace(' ', '_')
    for key, aliases in HEADER_ALIASES.items():
        normalized_aliases = {_norm(a) for a in aliases} | {str(a).replace(' ', '_').lower() for a in aliases}
        if normalized in normalized_aliases or compact in normalized_aliases:
            return key
    return None


def _read_csv(file_storage):
    raw = file_storage.read()
    text = raw.decode('utf-8-sig', errors='replace')
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t,')
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ';'
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return list(reader)


def _read_xlsx(file_storage):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError('Pro import XLSX je potřeba přidat openpyxl do requirements.txt.') from exc

    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or '').strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        if not row or all(cell in (None, '') for cell in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))})
    return data


def read_supplier_rows(file_storage):
    filename = (file_storage.filename or '').lower()
    if filename.endswith('.xlsx'):
        rows = _read_xlsx(file_storage)
    elif filename.endswith('.csv') or filename.endswith('.txt'):
        rows = _read_csv(file_storage)
    else:
        raise ValueError('Nahraj prosím CSV nebo XLSX soubor.')

    normalized_rows = []
    for row in rows:
        out = {}
        for header, value in row.items():
            key = _header_key(header)
            if key:
                out[key] = '' if value is None else str(value).strip()
        if any(out.values()):
            normalized_rows.append(out)
    return normalized_rows


def _find_variant(row):
    internal_sku = (row.get('internal_sku') or '').strip()
    if internal_sku:
        variant = ProductVariant.query.filter_by(sku=internal_sku).first()
        if variant:
            return variant

    product_name = row.get('product_name') or ''
    size = row.get('size') or ''
    color = row.get('color') or row.get('supplier_color') or ''

    if not product_name or not size:
        return None

    product_key = _norm(product_name)
    products = Product.query.filter(Product.active == True).all()
    matched_products = []
    for product in products:
        name_key = _norm(product.name)
        if product_key == name_key or product_key in name_key or name_key in product_key:
            matched_products.append(product)

    size_key = _compact(size)
    color_key = _color_key(color)
    for product in matched_products:
        for variant in product.variants:
            if _compact(variant.size) != size_key:
                continue
            if color and _color_key(variant.color) != color_key:
                continue
            return variant
    return None


def import_supplier_sku_file(file_storage, update_woocommerce=False):
    rows = read_supplier_rows(file_storage)
    result = {
        'total': len(rows),
        'matched': 0,
        'updated_wc': 0,
        'wc_errors': 0,
        'unmatched': [],
        'missing_supplier_sku': [],
        'updated': [],
    }

    for index, row in enumerate(rows, start=2):
        supplier_sku = (row.get('supplier_sku') or '').strip()
        if not supplier_sku:
            result['missing_supplier_sku'].append({'row': index, 'row_data': row})
            continue

        variant = _find_variant(row)
        if not variant:
            result['unmatched'].append({'row': index, 'row_data': row})
            continue

        variant.supplier_sku = supplier_sku
        variant.supplier_color = (row.get('supplier_color') or row.get('color') or '').strip()
        variant.supplier_product_code = (row.get('supplier_product_code') or '').strip()
        variant.supplier_ean = (row.get('supplier_ean') or '').strip()
        variant.supplier_sync_status = 'importováno'
        variant.supplier_sync_message = 'Dodavatelské SKU bylo importováno.'
        variant.supplier_synced_at = datetime.now()
        result['matched'] += 1
        result['updated'].append({
            'product': variant.product.name,
            'size': variant.size,
            'color': variant.color or '',
            'internal_sku': variant.sku,
            'supplier_sku': variant.supplier_sku,
            'supplier_color': variant.supplier_color,
        })

        if update_woocommerce:
            db.session.flush()
            wc_result = update_supplier_variant_in_woocommerce(variant)
            variant.supplier_sync_status = 'wc aktualizováno' if wc_result.get('ok') else 'wc chyba'
            variant.supplier_sync_message = wc_result.get('message', '')
            variant.supplier_synced_at = datetime.now()
            if wc_result.get('ok'):
                result['updated_wc'] += 1
            else:
                result['wc_errors'] += 1

    return result
